import csv
import json
import math
import re
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "DESCRIPTIVE LITHOLOGY CTSJ-30 (P-02) Running.xlsx"
OUT_DIR = ROOT / "sample-data" / "ctsj-30"


LITHOLOGY_ALIASES = {
    "CARB SHALE": ("CARBSHL", "Carbonaceous Shale", "#4b5563"),
    "SH COAL": ("SHCOAL", "Shaly Coal", "#30261f"),
    "SHALY SANDSTONE": ("SHSST", "Shaly Sandstone", "#b79a64"),
    "SANDY SHALE": ("SANDYSH", "Sandy Shale", "#8c8577"),
    "SANDY SOIL": ("SOILSNDY", "Sandy Soil", "#c9a36a"),
    "SANDSTONE": ("SST", "Sandstone", "#c9975b"),
    "SHALE": ("SHALE", "Shale", "#6b7280"),
    "COAL": ("COAL", "Coal", "#202124"),
    "SOIL": ("SOIL", "Soil", "#b8a36a"),
    "SLUDGE": ("SLUDGE", "Sludge", "#7a8875"),
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and math.isclose(value, round(value), abs_tol=1e-9):
        return int(round(value))
    return value


def to_float(value):
    value = clean(value)
    if value is None:
        return None
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return None


def normalize_lithology(source):
    text = clean(source)
    if not text:
        return {"code": None, "label": "Unknown", "color": "#94a3b8", "review": True}
    key = str(text).upper()
    if key in LITHOLOGY_ALIASES:
        code, label, color = LITHOLOGY_ALIASES[key]
        return {"code": code, "label": label, "color": color, "review": False}
    if "COAL" in key:
        return {"code": key, "label": str(text).title(), "color": "#352820", "review": True}
    if "SH" in key:
        return {"code": key, "label": str(text).title(), "color": "#7c8794", "review": True}
    if "SST" in key or "SAND" in key:
        return {"code": key, "label": str(text).title(), "color": "#c9975b", "review": True}
    return {"code": key, "label": str(text).title(), "color": "#8aa29e", "review": True}


def find_depth_text(worksheet):
    values = []
    for row in worksheet.iter_rows():
        for cell in row:
            value = clean(cell.value)
            if isinstance(value, str) and any(token in value.upper() for token in ["RUNNING", "CLOSED", "DEPTH"]):
                values.append({"cell": cell.coordinate, "text": value})
    return values


def extract_clean():
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=False)
    worksheet = workbook["Sheet1"]
    depth_text = find_depth_text(worksheet)

    total_depth = None
    status = "unknown"
    for item in depth_text:
        text = item["text"].upper()
        if "CLOSED" in text:
            status = "closed"
            match = re.search(r"(\d+(?:\.\d+)?)\s*M", text)
            if match:
                total_depth = float(match.group(1))
    if total_depth is None:
        for item in depth_text:
            text = item["text"].upper()
            if "RUNNING" in text:
                status = "running"
                match = re.search(r"(\d+(?:\.\d+)?)\s*M", text)
                if match:
                    total_depth = float(match.group(1))

    intervals = []
    run_intervals = []
    remarks = []
    seam_intervals = []
    dictionary_review = []

    for row_number in range(9, 595):
        run_from = to_float(worksheet.cell(row_number, 1).value)
        run_to = to_float(worksheet.cell(row_number, 2).value)
        if run_from is not None and run_to is not None:
            run_intervals.append(
                {
                    "id": f"run-{row_number}",
                    "sourceRow": row_number,
                    "fromDepth": run_from,
                    "toDepth": run_to,
                    "thickness": to_float(worksheet.cell(row_number, 3).value),
                    "recovery": to_float(worksheet.cell(row_number, 4).value),
                    "recoveryPercent": to_float(worksheet.cell(row_number, 5).value),
                }
            )

        lith_from = to_float(worksheet.cell(row_number, 6).value)
        lith_thickness = to_float(worksheet.cell(row_number, 7).value)
        lithology = clean(worksheet.cell(row_number, 9).value)
        if lith_from is None or lith_thickness is None or lith_thickness <= 0 or not lithology:
            continue

        normalized = normalize_lithology(lithology)
        if normalized["review"]:
            dictionary_review.append({"sourceRow": row_number, "sourceText": lithology, "suggestedCode": normalized["code"]})

        interval = {
            "id": f"ctsj-lith-{row_number}",
            "sourceRow": row_number,
            "fromDepth": lith_from,
            "toDepth": round(lith_from + lith_thickness, 3),
            "thickness": lith_thickness,
            "recovery": to_float(worksheet.cell(row_number, 8).value),
            "lithologySource": lithology,
            "lithologyCode": normalized["code"],
            "lithologyLabel": normalized["label"],
            "displayColor": normalized["color"],
            "grainSize": clean(worksheet.cell(row_number, 10).value),
            "loggedColor": clean(worksheet.cell(row_number, 11).value),
            "rqdPieceLengths": clean(worksheet.cell(row_number, 12).value),
            "rqdPercent": to_float(worksheet.cell(row_number, 13).value),
            "structuralFeatures": clean(worksheet.cell(row_number, 14).value),
            "coreDip": clean(worksheet.cell(row_number, 15).value),
            "seamName": clean(worksheet.cell(row_number, 16).value),
            "remark": clean(worksheet.cell(row_number, 17).value),
        }
        intervals.append(interval)

        if interval["structuralFeatures"] or interval["remark"]:
            remarks.append(
                {
                    "id": f"ctsj-remark-{row_number}",
                    "sourceRow": row_number,
                    "depth": interval["fromDepth"],
                    "toDepth": interval["toDepth"],
                    "text": " / ".join(
                        part for part in [interval["structuralFeatures"], interval["remark"]] if part
                    ),
                }
            )

        if interval["seamName"]:
            seam_intervals.append(
                {
                    "id": f"ctsj-seam-{row_number}",
                    "sourceRow": row_number,
                    "name": interval["seamName"],
                    "fromDepth": interval["fromDepth"],
                    "toDepth": interval["toDepth"],
                    "thickness": interval["thickness"],
                    "lithologyCode": interval["lithologyCode"],
                    "lithologyLabel": interval["lithologyLabel"],
                }
            )

    return {
        "borehole": {
            "id": "CTSJ-30-P02",
            "title": "CTSJ-30(P-02) Descriptive Lithology",
            "block": "SARADHAPUR-JALATAP",
            "status": status,
            "totalDepth": total_depth,
            "sourceWorkbook": WORKBOOK.name,
            "sourceSheet": "Sheet1",
            "sourceDepthText": depth_text,
        },
        "summary": {
            "runIntervalCount": len(run_intervals),
            "lithologyIntervalCount": len(intervals),
            "seamIntervalCount": len(seam_intervals),
            "remarkCount": len(remarks),
            "dictionaryReviewCount": len(dictionary_review),
        },
        "runIntervals": run_intervals,
        "lithologyIntervals": intervals,
        "seamIntervals": seam_intervals,
        "remarks": remarks,
        "dictionaryReview": dictionary_review,
    }


def lithology_at(intervals, depth):
    for interval in intervals:
        if interval["fromDepth"] <= depth <= interval["toDepth"]:
            return interval
    return intervals[-1]


def base_curve_values(interval, depth):
    code = interval["lithologyCode"] or ""
    wiggle = math.sin(depth * 0.33) * 3 + math.sin(depth * 0.09) * 2
    if "COAL" in code:
        return 38 + wiggle, 1.34 + math.sin(depth * 0.21) * 0.03, 120 + math.sin(depth * 0.19) * 14
    if "SHALE" in code or "SH" in code:
        return 94 + wiggle, 2.28 + math.sin(depth * 0.17) * 0.05, 42 + math.sin(depth * 0.13) * 8
    if "SST" in code or "SAND" in code:
        return 67 + wiggle, 2.42 + math.sin(depth * 0.11) * 0.06, 78 + math.sin(depth * 0.15) * 12
    return 72 + wiggle, 2.05 + math.sin(depth * 0.1) * 0.05, 55 + math.sin(depth * 0.12) * 10


def build_curves(intervals):
    rows = []
    depth = 0.0
    max_depth = intervals[-1]["toDepth"]
    while depth <= max_depth + 0.001:
        interval = lithology_at(intervals, depth)
        gamma, density, resistivity = base_curve_values(interval, depth)

        # Known planted curve anomaly: coal-like log response inside a sandstone interval.
        if 82 <= depth <= 86:
            gamma = 35 + math.sin(depth) * 2
            density = 1.42 + math.sin(depth * 0.4) * 0.02
            resistivity = 135 + math.sin(depth * 0.4) * 12

        # Known planted curve anomaly: shale-like response inside a coal interval.
        if 334 <= depth <= 338:
            gamma = 118 + math.sin(depth) * 2
            density = 2.45 + math.sin(depth * 0.4) * 0.02
            resistivity = 32 + math.sin(depth * 0.4) * 4

        rows.append(
            {
                "depth": round(depth, 2),
                "gamma_api": round(gamma, 3),
                "density_gcc": round(density, 3),
                "resistivity_ohmm": round(resistivity, 3),
                "lithology_code": interval["lithologyCode"],
                "lithology_source": interval["lithologySource"],
            }
        )
        depth = round(depth + 0.5, 3)
    return rows


def build_ai_test(clean_dataset):
    dataset = deepcopy(clean_dataset)
    intervals = dataset["lithologyIntervals"]
    planted = []

    def mark(index, issue_type, description, before, after):
        interval = intervals[index]
        planted.append(
            {
                "issueType": issue_type,
                "intervalId": interval["id"],
                "sourceRow": interval["sourceRow"],
                "description": description,
                "before": before,
                "after": after,
            }
        )

    before = intervals[45]["lithologyCode"]
    intervals[45]["lithologyCode"] = None
    intervals[45]["lithologyLabel"] = "Unknown"
    mark(45, "missing_lithology_code", "Lithology code removed to test required-field validation.", before, None)

    before = intervals[88]["fromDepth"]
    intervals[88]["fromDepth"] = round(intervals[88]["fromDepth"] - 0.35, 3)
    mark(88, "overlap", "From-depth shifted upward to overlap previous interval.", before, intervals[88]["fromDepth"])

    before = intervals[132]["fromDepth"]
    intervals[132]["fromDepth"] = round(intervals[132]["fromDepth"] + 0.45, 3)
    mark(132, "gap", "From-depth shifted downward to create a gap after previous interval.", before, intervals[132]["fromDepth"])

    before = intervals[180]["recovery"]
    intervals[180]["recovery"] = round(intervals[180]["thickness"] + 0.35, 3)
    mark(180, "recovery_greater_than_thickness", "Recovery made larger than interval thickness.", before, intervals[180]["recovery"])

    before = intervals[230]["rqdPercent"]
    intervals[230]["rqdPercent"] = 128
    mark(230, "invalid_rqd_percent", "RQD percent set above 100.", before, 128)

    before = {"thickness": intervals[270]["thickness"], "toDepth": intervals[270]["toDepth"]}
    intervals[270]["thickness"] = -0.2
    intervals[270]["toDepth"] = round(intervals[270]["fromDepth"] - 0.2, 3)
    mark(
        270,
        "invalid_depth_range",
        "Thickness made negative so to-depth is less than from-depth.",
        before,
        {"thickness": intervals[270]["thickness"], "toDepth": intervals[270]["toDepth"]},
    )

    dataset["experiment"] = {
        "purpose": "Known-bad fixture for validation and AI-assisted correction demos.",
        "plantedIssues": planted,
        "curveAnomalies": [
            {
                "depthFrom": 82,
                "depthTo": 86,
                "description": "Coal-like low gamma/density response planted in a sandstone interval.",
            },
            {
                "depthFrom": 334,
                "depthTo": 338,
                "description": "Shale-like high gamma/density response planted in a coal interval.",
            },
        ],
    }
    dataset["summary"]["plantedIssueCount"] = len(planted)
    return dataset


def write_outputs():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    clean_dataset = extract_clean()
    curves = build_curves(clean_dataset["lithologyIntervals"])
    ai_dataset = build_ai_test(clean_dataset)

    (OUT_DIR / "ctsj-30-clean.json").write_text(json.dumps(clean_dataset, indent=2), encoding="utf-8")
    (OUT_DIR / "ctsj-30-ai-test.json").write_text(json.dumps(ai_dataset, indent=2), encoding="utf-8")
    (OUT_DIR / "ctsj-30-experiment-manifest.json").write_text(
        json.dumps(ai_dataset["experiment"], indent=2), encoding="utf-8"
    )

    with (OUT_DIR / "ctsj-30-synthetic-curves.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["depth", "gamma_api", "density_gcc", "resistivity_ohmm", "lithology_code", "lithology_source"],
        )
        writer.writeheader()
        writer.writerows(curves)

    print(json.dumps({"clean": clean_dataset["summary"], "aiTest": ai_dataset["summary"], "curves": len(curves)}, indent=2))


if __name__ == "__main__":
    write_outputs()
