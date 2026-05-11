import json
import math
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "DOC-20260510-WA0000..xlsx"
OUT_JSON = ROOT / "sample-data" / "pbh-62-normalized.json"
OUT_JS = ROOT / "sample-data" / "pbh-62-normalized.js"


LITHOLOGY_GROUPS = {
    "COAL": ("Coal", "#202124"),
    "SHCOAL": ("Shaly Coal", "#3a3029"),
    "CARBSHL": ("Carbonaceous Shale", "#4b5563"),
    "CARBSHH": ("Carbonaceous Shale", "#374151"),
    "CARBCLAY": ("Carbonaceous Clay", "#5f6368"),
    "SHALE": ("Shale", "#6b7280"),
    "GREYSH": ("Grey Shale", "#7c8794"),
    "SANDYSH": ("Sandy Shale", "#8c8577"),
    "SHSST": ("Shaly Sandstone", "#b79a64"),
    "CLAY": ("Clay", "#a1784a"),
    "CLAYSTONE": ("Claystone", "#9b7653"),
    "DOLERITE": ("Dolerite", "#37515f"),
    "SOILSNDY": ("Sandy Soil", "#c9a36a"),
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and math.isclose(value, round(value), abs_tol=1e-9):
        return int(round(value))
    return value


def to_float(value):
    value = clean(value)
    if value is None:
        return None
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return None


def normalize_lithology(code):
    code = clean(code)
    if not code:
        return {"code": None, "label": "Unknown", "color": "#94a3b8"}
    key = str(code).upper()
    if key in LITHOLOGY_GROUPS:
        label, color = LITHOLOGY_GROUPS[key]
        return {"code": key, "label": label, "color": color}
    if key.startswith("SS"):
        return {"code": key, "label": "Sandstone", "color": "#c9975b"}
    if "SH" in key:
        return {"code": key, "label": "Shale/Sandstone", "color": "#9a9484"}
    return {"code": key, "label": key, "color": "#8aa29e"}


def box_for_depth(depth, total_depth, box_count):
    if depth is None or not total_depth or not box_count:
        return None
    box = int(math.ceil((float(depth) / float(total_depth)) * box_count))
    return max(1, min(box_count, box))


def load_core_images():
    folder = ROOT / "MTSE-65(PBH 62)"
    images = []
    if not folder.exists():
        return images
    for path in sorted(folder.glob("*.jpg"), key=lambda p: int(p.stem) if p.stem.isdigit() else p.stem):
        images.append(
            {
                "boxNumber": int(path.stem) if path.stem.isdigit() else path.stem,
                "file": f"../MTSE-65(PBH 62)/{path.name}",
                "name": path.name,
            }
        )
    return images


def extract():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    worksheet = workbook["abcd"]

    total_depth = 604.3
    closure_note = None
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and "BH CLOSED" in value.upper():
                closure_note = value.strip()
                match = re.search(r"(\d+(?:\.\d+)?)\s*M", value.upper())
                if match:
                    total_depth = float(match.group(1))

    core_images = load_core_images()
    box_count = len(core_images)
    run_intervals = []
    lithology_intervals = []
    seam_intervals = []
    remarks = []
    validation_issues = []

    previous_to = None
    for row_number, row in enumerate(worksheet.iter_rows(min_row=12, values_only=True), start=12):
        row = list(row) + [None] * 29

        run_from = to_float(row[0])
        run_to = to_float(row[1])
        run_thickness = to_float(row[2])
        run_recovery = to_float(row[3])

        lith_from = to_float(row[4])
        lith_thickness = to_float(row[5])
        lith_recovery = to_float(row[6])
        lithology = clean(row[7])
        color = clean(row[8])
        structural_features = clean(row[9])
        core_dip = clean(row[10])
        seam_name = clean(row[11])
        rqd = to_float(row[12])
        remark = clean(row[13])

        if run_from is not None and run_to is not None:
            issue = None
            if run_to <= run_from:
                issue = "Run to-depth is not greater than from-depth."
            if run_recovery is not None and run_thickness is not None and run_recovery > run_thickness + 0.001:
                issue = "Run recovery is greater than run thickness."
            run_intervals.append(
                {
                    "id": f"run-{row_number}",
                    "sourceRow": row_number,
                    "fromDepth": run_from,
                    "toDepth": run_to,
                    "thickness": run_thickness,
                    "recovery": run_recovery,
                    "recoveryPercent": round((run_recovery / run_thickness) * 100, 1)
                    if run_recovery is not None and run_thickness
                    else None,
                    "issue": issue,
                }
            )
            if issue:
                validation_issues.append({"sourceRow": row_number, "severity": "warning", "message": issue})

        if lith_from is None or lithology is None:
            continue

        lith_to = round(lith_from + lith_thickness, 3) if lith_thickness is not None else None
        normalized = normalize_lithology(lithology)
        image_box = box_for_depth(lith_from, total_depth, box_count)
        image = core_images[image_box - 1] if image_box and image_box <= len(core_images) else None

        interval_issue = None
        if lith_to is None or lith_to <= lith_from:
            interval_issue = "Lithology to-depth is not greater than from-depth."
        elif previous_to is not None and lith_from < previous_to - 0.001:
            interval_issue = "Lithology interval overlaps the previous interval."
        elif previous_to is not None and lith_from > previous_to + 0.001:
            interval_issue = "Gap before this lithology interval."

        interval = {
            "id": f"lith-{row_number}",
            "sourceRow": row_number,
            "fromDepth": lith_from,
            "toDepth": lith_to,
            "thickness": lith_thickness,
            "recovery": lith_recovery,
            "recoveryPercent": round((lith_recovery / lith_thickness) * 100, 1)
            if lith_recovery is not None and lith_thickness
            else None,
            "lithologyCode": normalized["code"],
            "lithologyLabel": normalized["label"],
            "displayColor": normalized["color"],
            "loggedColor": color,
            "structuralFeatures": structural_features,
            "coreDip": core_dip,
            "seamName": seam_name,
            "rqd": rqd,
            "remark": remark,
            "imageBox": image_box,
            "imageFile": image["file"] if image else None,
            "issue": interval_issue,
        }
        lithology_intervals.append(interval)

        if interval_issue:
            validation_issues.append({"sourceRow": row_number, "severity": "warning", "message": interval_issue})

        if structural_features or remark:
            remarks.append(
                {
                    "id": f"remark-{row_number}",
                    "sourceRow": row_number,
                    "depth": lith_from,
                    "toDepth": lith_to,
                    "text": " / ".join([part for part in [structural_features, remark] if part]),
                    "kind": "structure" if structural_features else "remark",
                    "imageBox": image_box,
                }
            )

        if seam_name:
            seam_intervals.append(
                {
                    "id": f"seam-{row_number}",
                    "sourceRow": row_number,
                    "name": str(seam_name),
                    "fromDepth": lith_from,
                    "toDepth": lith_to,
                    "thickness": lith_thickness,
                    "lithologyCode": normalized["code"],
                    "lithologyLabel": normalized["label"],
                    "imageBox": image_box,
                }
            )

        previous_to = lith_to if lith_to is not None else previous_to

    lithology_counts = {}
    for interval in lithology_intervals:
        code = interval["lithologyCode"] or "UNKNOWN"
        lithology_counts[code] = lithology_counts.get(code, 0) + 1

    dataset = {
        "borehole": {
            "id": "PBH-62",
            "projectCode": "MTSE-65",
            "title": "MTSE-65 / PBH-62 Coal Borehole",
            "state": "Madhya Pradesh",
            "totalDepth": total_depth,
            "closureNote": closure_note,
            "sourceWorkbook": WORKBOOK.name,
            "sourceSheet": "abcd",
        },
        "summary": {
            "runIntervalCount": len(run_intervals),
            "lithologyIntervalCount": len(lithology_intervals),
            "seamIntervalCount": len(seam_intervals),
            "remarkCount": len(remarks),
            "coreImageCount": len(core_images),
            "validationIssueCount": len(validation_issues),
            "lithologyCounts": dict(sorted(lithology_counts.items(), key=lambda item: item[0])),
        },
        "runIntervals": run_intervals,
        "lithologyIntervals": lithology_intervals,
        "seamIntervals": seam_intervals,
        "remarks": remarks,
        "coreImages": core_images,
        "validationIssues": validation_issues,
        "prototypeNotes": {
            "coreImageDepthLinking": "Core image depth links are approximate, based on ordered box photographs distributed across total borehole depth. Future work should capture exact box depth ranges.",
            "curveTracks": "Gamma and density curves in the prototype are generated demo curves until LAS/geophysical data is supplied.",
        },
    }
    return dataset


def main():
    dataset = extract()
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(dataset, indent=2), encoding="utf-8")
    OUT_JS.write_text(
        "window.PBH62_DATA = " + json.dumps(dataset, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT_JSON.relative_to(ROOT)}")
    print(f"Wrote {OUT_JS.relative_to(ROOT)}")
    print(json.dumps(dataset["summary"], indent=2))


if __name__ == "__main__":
    main()
