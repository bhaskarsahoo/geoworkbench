from pathlib import Path

from openpyxl import Workbook


def write_ctsj_workbook(path: Path, code: str, block: str, depth_shift: float = 0.0) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Descriptive Lithology"
    sheet["A1"] = "DEMO IMPORT WORKBOOK"
    sheet["A3"] = "BOREHOLE"
    sheet["B3"] = code
    sheet["A4"] = "BLOCK"
    sheet["B4"] = block
    sheet["A7"] = "DRILLING RUN"
    sheet["F7"] = "DEPTH & THICKNESS AFTER ADJUSTMENT"
    headers = [
        "Run From",
        "Run To",
        "Run Thickness",
        "Run Recovery",
        "Run Recovery %",
        "Lith From",
        "Lith Thickness",
        "Lith Recovery",
        "Lithology",
        "Grain Size",
        "Colour",
        "RQD Pieces",
        "RQD %",
        "Structural Features",
        "Core Dip",
        "Seam",
        "Remarks",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(8, column).value = header

    rows = [
        (0, 24, 22.8, "SOIL", "", "Brown", 45, "", "", "Weathered zone"),
        (24, 58, 31.1, "SANDSTONE", "Fine", "Grey", 68, "", "", "Competent sandstone"),
        (58, 93, 29.4, "SHALE", "", "Dark grey", 52, "Inclined fractures", "", "Fractured shale"),
        (93, 107, 11.6, "COAL", "", "Black", 44, "Cleat visible", "Seam A", "Coal seam picked at site"),
        (107, 148, 36.7, "CARB SHALE", "", "Blackish grey", 49, "Laminated", "", "Carbonaceous shale"),
        (148, 166, 13.1, "COAL", "", "Black", 39, "", "Seam B", "Corrected seam interval"),
        (166, 186, 17.5, "SANDSTONE", "Medium", "Light grey", 71, "", "", "Base sandstone"),
    ]
    for index, (from_depth, to_depth, recovery, lithology, grain, color, rqd, features, seam, remarks) in enumerate(
        rows,
        start=9,
    ):
        from_depth = round(from_depth + depth_shift, 2)
        to_depth = round(to_depth + depth_shift, 2)
        thickness = round(to_depth - from_depth, 2)
        sheet.cell(index, 1).value = from_depth
        sheet.cell(index, 2).value = to_depth
        sheet.cell(index, 3).value = thickness
        sheet.cell(index, 4).value = recovery
        sheet.cell(index, 5).value = round(recovery / thickness * 100, 1)
        sheet.cell(index, 6).value = from_depth
        sheet.cell(index, 7).value = thickness
        sheet.cell(index, 8).value = recovery
        sheet.cell(index, 9).value = lithology
        sheet.cell(index, 10).value = grain
        sheet.cell(index, 11).value = color
        sheet.cell(index, 12).value = ""
        sheet.cell(index, 13).value = rqd
        sheet.cell(index, 14).value = features
        sheet.cell(index, 15).value = ""
        sheet.cell(index, 16).value = seam
        sheet.cell(index, 17).value = remarks

    sheet.cell(17, 1).value = f"BOREHOLE CLOSED AT {186 + depth_shift:.1f} M"
    workbook.save(path)


def write_curve_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["depth_m,gamma_api,resistivity_ohmm,density_gcc,source_note"]
    depth = 0
    while depth <= 186:
        gamma = 42 if 93 <= depth <= 107 or 148 <= depth <= 166 else 86
        resistivity = 92 if 93 <= depth <= 107 or 148 <= depth <= 166 else 34
        density = 1.52 if 93 <= depth <= 107 or 148 <= depth <= 166 else 2.34
        lines.append(f"{depth},{gamma + (depth % 7)},{resistivity + (depth % 5)},{density:.2f},demo curve mapping")
        depth += 2
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    repo_root = Path(__file__).resolve().parents[2]
    output = repo_root / "sample-data" / "import-demo"
    write_ctsj_workbook(output / "IMPORT-DEMO-01-CTSJ-template.xlsx", "IMPORT-DEMO-01", "Demo Import Block")
    write_ctsj_workbook(
        output / "IMPORT-DEMO-02-corrected-section.xlsx",
        "IMPORT-DEMO-02",
        "Demo Corrected Section",
        depth_shift=4.5,
    )
    write_curve_csv(output / "IMPORT-DEMO-03-curve-samples.csv")
    print(f"Wrote import demo fixtures to {output}")


if __name__ == "__main__":
    main()
