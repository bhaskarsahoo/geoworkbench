import math
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import (
    Borehole,
    Curve,
    CurveSample,
    DisplayLayout,
    LithologyInterval,
    Project,
    SeamInterval,
    Site,
    SourceImport,
)
from app.domains.display_layouts.defaults import default_borehole_layout
from app.services.validation.borehole_validation import replace_validation_issues, validate_borehole


@dataclass(frozen=True)
class ExcelTemplate:
    key: str
    label: str
    sheet_name: str
    header_rows: list[int]
    data_start_row: int
    data_end_row: int | None
    columns: dict[str, int]


LITHOLOGY_ALIASES = {
    "CARB SHALE": ("CARBSHL", "Carbonaceous Shale", "#4b5563"),
    "CARBSHL": ("CARBSHL", "Carbonaceous Shale", "#4b5563"),
    "CARBSHH": ("CARBSHH", "Carbonaceous Shale", "#374151"),
    "CARBCLAY": ("CARBCLAY", "Carbonaceous Clay", "#5f6368"),
    "CLAY": ("CLAY", "Clay", "#a1784a"),
    "CLAYSTONE": ("CLAYSTONE", "Claystone", "#9b7653"),
    "COAL": ("COAL", "Coal", "#202124"),
    "DOLERITE": ("DOLERITE", "Dolerite", "#37515f"),
    "GREYSH": ("GREYSH", "Grey Shale", "#7c8794"),
    "SANDSTONE": ("SST", "Sandstone", "#c9975b"),
    "SANDY SHALE": ("SANDYSH", "Sandy Shale", "#8c8577"),
    "SANDY SOIL": ("SOILSNDY", "Sandy Soil", "#c9a36a"),
    "SHALE": ("SHALE", "Shale", "#6b7280"),
    "SH COAL": ("SHCOAL", "Shaly Coal", "#30261f"),
    "SHALY SANDSTONE": ("SHSST", "Shaly Sandstone", "#b79a64"),
    "SHCOAL": ("SHCOAL", "Shaly Coal", "#3a3029"),
    "SHSST": ("SHSST", "Shaly Sandstone", "#b79a64"),
    "SLUDGE": ("SLUDGE", "Sludge", "#7a8875"),
    "SOIL": ("SOIL", "Soil", "#b8a36a"),
    "SOILSNDY": ("SOILSNDY", "Sandy Soil", "#c9a36a"),
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and math.isclose(value, round(value), abs_tol=1e-9):
        return int(round(value))
    return value


def to_float(value) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return None


def cell_text(worksheet: Worksheet) -> list[dict]:
    values = []
    for row in worksheet.iter_rows():
        for cell in row:
            value = clean(cell.value)
            if value is not None:
                values.append({"cell": cell.coordinate, "text": str(value)})
    return values


def detect_template(workbook_path: Path) -> tuple[ExcelTemplate, list[dict], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    warnings: list[str] = []
    candidates = []
    for worksheet in workbook.worksheets:
        text = " ".join(item["text"].upper() for item in cell_text(worksheet)[:250])
        if "LITHOLOGY DEPTH" in text and "DESCRIPTION AS PER CORE RECOVERY" in text:
            candidates.append(
                ExcelTemplate(
                    key="pbh_descriptive_v1",
                    label="PBH descriptive lithology workbook",
                    sheet_name=worksheet.title,
                    header_rows=[9, 10, 11],
                    data_start_row=12,
                    data_end_row=None,
                    columns={
                        "run_from": 1,
                        "run_to": 2,
                        "run_thickness": 3,
                        "run_recovery": 4,
                        "lith_from": 5,
                        "lith_thickness": 6,
                        "lith_recovery": 7,
                        "lithology": 8,
                        "logged_color": 9,
                        "features": 10,
                        "core_dip": 11,
                        "seam": 12,
                        "rqd": 13,
                        "remarks": 14,
                    },
                )
            )
        if "DEPTH & THICKNESS AFTER ADJUSTMENT" in text and "DRILLING RUN" in text:
            candidates.append(
                ExcelTemplate(
                    key="ctsj_descriptive_v1",
                    label="CTSJ descriptive lithology workbook",
                    sheet_name=worksheet.title,
                    header_rows=[7, 8],
                    data_start_row=9,
                    data_end_row=None,
                    columns={
                        "run_from": 1,
                        "run_to": 2,
                        "run_thickness": 3,
                        "run_recovery": 4,
                        "run_recovery_percent": 5,
                        "lith_from": 6,
                        "lith_thickness": 7,
                        "lith_recovery": 8,
                        "lithology": 9,
                        "grain_size": 10,
                        "logged_color": 11,
                        "rqd_piece_lengths": 12,
                        "rqd": 13,
                        "features": 14,
                        "core_dip": 15,
                        "seam": 16,
                        "remarks": 17,
                    },
                )
            )

    if not candidates:
        raise ValueError("No supported descriptive lithology template was detected.")
    if len(candidates) > 1:
        warnings.append("Multiple possible templates detected; using the first highest-confidence match.")
    template = candidates[0]
    return template, [{"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column} for sheet in workbook.worksheets], warnings


def normalize_lithology(value) -> dict:
    source = clean(value)
    if not source:
        return {"code": "UNKNOWN", "label": "Unknown", "color": "#94a3b8", "needs_review": True}
    key = str(source).upper()
    if key in LITHOLOGY_ALIASES:
        code, label, color = LITHOLOGY_ALIASES[key]
        return {"code": code, "label": label, "color": color, "needs_review": False}
    if key.startswith("SS") or "SST" in key or "SAND" in key:
        return {"code": key, "label": str(source).title(), "color": "#c9975b", "needs_review": True}
    if "COAL" in key:
        return {"code": key, "label": str(source).title(), "color": "#30261f", "needs_review": True}
    if "SH" in key:
        return {"code": key, "label": str(source).title(), "color": "#7c8794", "needs_review": True}
    return {"code": key, "label": str(source).title(), "color": "#8aa29e", "needs_review": True}


def extract_metadata(worksheet: Worksheet, template: ExcelTemplate) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    text_values = cell_text(worksheet)
    metadata = {
        "source_sheet": worksheet.title,
        "source_depth_text": [],
        "borehole_code": None,
        "block": None,
        "state": None,
        "total_depth": None,
        "closure_note": None,
        "workflow_status": "imported_for_review",
    }

    for item in text_values:
        text = item["text"].strip()
        upper = text.upper()
        if any(token in upper for token in ["RUNNING", "CLOSED", "TOTAL DEPTH", "BOREHOLE DEPTH"]):
            metadata["source_depth_text"].append(item)
        if "BH CLOSED" in upper or "BOERHOLE CLOSED" in upper or "BOREHOLE CLOSED" in upper:
            metadata["closure_note"] = text
            metadata["workflow_status"] = "ready_for_central_review"
            match = re.search(r"(\d+(?:\.\d+)?)\s*M", upper)
            if match:
                metadata["total_depth"] = float(match.group(1))
        elif "RUNNING" in upper and metadata["total_depth"] is None:
            metadata["workflow_status"] = "logging_in_progress"
            match = re.search(r"(\d+(?:\.\d+)?)\s*M", upper)
            if match:
                metadata["total_depth"] = float(match.group(1))
        elif "TOTAL DEPTH" in upper and metadata["total_depth"] is None:
            match = re.search(r"(\d+(?:\.\d+)?)\s*M", upper)
            if match:
                metadata["total_depth"] = float(match.group(1))

    if template.key == "ctsj_descriptive_v1":
        metadata["borehole_code"] = clean(worksheet["B3"].value)
        metadata["block"] = clean(worksheet["B4"].value)
    if template.key == "pbh_descriptive_v1":
        metadata["state"] = clean(worksheet["L3"].value)
        metadata["borehole_code"] = "PBH-62"

    if not metadata["total_depth"]:
        warnings.append("Total depth could not be confidently detected; using last interval depth.")
    if len(metadata["source_depth_text"]) > 1:
        warnings.append("Multiple depth/status text values were found; review import diagnostics.")
    return metadata, warnings


def find_data_end_row(worksheet: Worksheet, template: ExcelTemplate) -> int:
    columns = template.columns
    last_data_row = template.data_start_row
    for row_number in range(template.data_start_row, worksheet.max_row + 1):
        row_text = " ".join(
            str(clean(worksheet.cell(row_number, col).value) or "").upper()
            for col in range(1, min(worksheet.max_column, 20) + 1)
        )
        if "BH CLOSED" in row_text or "BOERHOLE CLOSED" in row_text or "BOREHOLE CLOSED" in row_text:
            break
        lith_from = to_float(worksheet.cell(row_number, columns["lith_from"]).value)
        thickness = to_float(worksheet.cell(row_number, columns["lith_thickness"]).value)
        lithology = clean(worksheet.cell(row_number, columns["lithology"]).value)
        if lith_from is not None and thickness is not None and thickness != 0 and lithology:
            last_data_row = row_number
    return last_data_row


def rqd_to_fraction(value, template: ExcelTemplate) -> float | None:
    rqd = to_float(value)
    if rqd is None:
        return None
    if template.key == "ctsj_descriptive_v1":
        return round(rqd / 100, 4)
    return rqd


def profile_excel_workbook(workbook_path: Path) -> dict:
    template, sheets, warnings = detect_template(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    worksheet = workbook[template.sheet_name]
    metadata, metadata_warnings = extract_metadata(worksheet, template)
    end_row = find_data_end_row(worksheet, template)
    columns = template.columns

    run_count = 0
    lithology_count = 0
    dictionary_review: dict[str, int] = {}
    min_depth = None
    max_depth = None
    sample_rows = []
    for row_number in range(template.data_start_row, end_row + 1):
        if to_float(worksheet.cell(row_number, columns["run_from"]).value) is not None:
            run_count += 1
        lith_from = to_float(worksheet.cell(row_number, columns["lith_from"]).value)
        thickness = to_float(worksheet.cell(row_number, columns["lith_thickness"]).value)
        lithology = clean(worksheet.cell(row_number, columns["lithology"]).value)
        if lith_from is None or thickness is None or not lithology:
            continue
        normalized = normalize_lithology(lithology)
        if normalized["needs_review"]:
            dictionary_review[normalized["code"]] = dictionary_review.get(normalized["code"], 0) + 1
        lithology_count += 1
        min_depth = lith_from if min_depth is None else min(min_depth, lith_from)
        max_depth = lith_from + thickness if max_depth is None else max(max_depth, lith_from + thickness)
        if len(sample_rows) < 5:
            sample_rows.append(
                {
                    "source_row": row_number,
                    "from_depth": lith_from,
                    "to_depth": round(lith_from + thickness, 3),
                    "lithology_source": lithology,
                    "normalized_code": normalized["code"],
                }
            )

    return {
        "parser": "excel_profile",
        "workbook": workbook_path.name,
        "sheets": sheets,
        "template": {
            "key": template.key,
            "label": template.label,
            "sheet_name": template.sheet_name,
            "header_rows": template.header_rows,
            "data_start_row": template.data_start_row,
            "data_end_row": end_row,
            "columns": template.columns,
        },
        "metadata": metadata,
        "summary": {
            "run_interval_count": run_count,
            "lithology_interval_count": lithology_count,
            "min_depth": min_depth,
            "max_depth": round(max_depth, 3) if max_depth is not None else None,
            "dictionary_review_count": sum(dictionary_review.values()),
            "dictionary_review_codes": dictionary_review,
        },
        "sample_rows": sample_rows,
        "warnings": warnings + metadata_warnings,
    }


def normalize_excel_workbook(workbook_path: Path) -> dict:
    profile = profile_excel_workbook(workbook_path)
    template = ExcelTemplate(
        key=profile["template"]["key"],
        label=profile["template"]["label"],
        sheet_name=profile["template"]["sheet_name"],
        header_rows=profile["template"]["header_rows"],
        data_start_row=profile["template"]["data_start_row"],
        data_end_row=profile["template"]["data_end_row"],
        columns=profile["template"]["columns"],
    )
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    worksheet = workbook[template.sheet_name]
    columns = template.columns
    metadata = profile["metadata"]
    intervals = []
    seam_intervals = []
    run_intervals = []

    for row_number in range(template.data_start_row, (template.data_end_row or worksheet.max_row) + 1):
        run_from = to_float(worksheet.cell(row_number, columns["run_from"]).value)
        run_to = to_float(worksheet.cell(row_number, columns["run_to"]).value)
        if run_from is not None and run_to is not None:
            run_thickness = to_float(worksheet.cell(row_number, columns["run_thickness"]).value)
            run_recovery = to_float(worksheet.cell(row_number, columns["run_recovery"]).value)
            run_intervals.append(
                {
                    "id": f"run-{row_number}",
                    "sourceRow": row_number,
                    "fromDepth": run_from,
                    "toDepth": run_to,
                    "thickness": run_thickness,
                    "recovery": run_recovery,
                    "recoveryPercent": to_float(
                        worksheet.cell(row_number, columns.get("run_recovery_percent", -1)).value
                    )
                    if columns.get("run_recovery_percent")
                    else round((run_recovery / run_thickness) * 100, 1)
                    if run_recovery is not None and run_thickness
                    else None,
                }
            )

        lith_from = to_float(worksheet.cell(row_number, columns["lith_from"]).value)
        lith_thickness = to_float(worksheet.cell(row_number, columns["lith_thickness"]).value)
        lithology = clean(worksheet.cell(row_number, columns["lithology"]).value)
        if lith_from is None or lith_thickness is None or not lithology:
            continue
        normalized = normalize_lithology(lithology)
        lith_to = round(lith_from + lith_thickness, 3)
        seam_name = clean(worksheet.cell(row_number, columns["seam"]).value)
        interval = {
            "id": f"lith-{row_number}",
            "sourceRow": row_number,
            "fromDepth": lith_from,
            "toDepth": lith_to,
            "thickness": lith_thickness,
            "recovery": to_float(worksheet.cell(row_number, columns["lith_recovery"]).value),
            "lithologySource": lithology,
            "lithologyCode": normalized["code"],
            "lithologyLabel": normalized["label"],
            "displayColor": normalized["color"],
            "grainSize": clean(worksheet.cell(row_number, columns.get("grain_size", -1)).value)
            if columns.get("grain_size")
            else None,
            "loggedColor": clean(worksheet.cell(row_number, columns["logged_color"]).value),
            "structuralFeatures": clean(worksheet.cell(row_number, columns["features"]).value),
            "coreDip": clean(worksheet.cell(row_number, columns["core_dip"]).value),
            "seamName": seam_name,
            "rqd": rqd_to_fraction(worksheet.cell(row_number, columns["rqd"]).value, template),
            "rqdSource": clean(worksheet.cell(row_number, columns["rqd"]).value),
            "rqdPieceLengths": clean(worksheet.cell(row_number, columns.get("rqd_piece_lengths", -1)).value)
            if columns.get("rqd_piece_lengths")
            else None,
            "remark": clean(worksheet.cell(row_number, columns["remarks"]).value),
        }
        intervals.append(interval)
        if seam_name:
            seam_intervals.append(
                {
                    "id": f"seam-{row_number}",
                    "sourceRow": row_number,
                    "name": str(seam_name),
                    "fromDepth": lith_from,
                    "toDepth": lith_to,
                    "thickness": lith_thickness,
                    "lithologyCode": normalized["code"],
                    "lithologyLabel": normalized["label"],
                }
            )

    total_depth = metadata.get("total_depth") or (intervals[-1]["toDepth"] if intervals else 0)
    borehole_code = metadata.get("borehole_code") or workbook_path.stem
    safe_code = re.sub(r"[^A-Za-z0-9_-]+", "-", str(borehole_code)).strip("-").upper()
    return {
        "profile": profile,
        "borehole": {
            "id": safe_code,
            "title": f"{borehole_code} Descriptive Lithology",
            "block": metadata.get("block"),
            "state": metadata.get("state"),
            "totalDepth": total_depth,
            "closureNote": metadata.get("closure_note"),
            "workflowStatus": metadata.get("workflow_status"),
            "sourceWorkbook": workbook_path.name,
            "sourceSheet": template.sheet_name,
        },
        "runIntervals": run_intervals,
        "lithologyIntervals": intervals,
        "seamIntervals": seam_intervals,
    }


def synthetic_curve_value(curve_key: str, depth: float, lithology_code: str | None) -> float:
    code = lithology_code or ""
    wiggle = math.sin(depth * 0.33) * 3 + math.sin(depth * 0.09) * 2
    if "COAL" in code:
        base = {"gamma": 38, "density": 1.35, "resistivity": 120}[curve_key]
    elif "SHALE" in code or "SH" in code:
        base = {"gamma": 94, "density": 2.28, "resistivity": 42}[curve_key]
    elif "SST" in code or "SAND" in code or code.startswith("SS"):
        base = {"gamma": 67, "density": 2.42, "resistivity": 78}[curve_key]
    else:
        base = {"gamma": 72, "density": 2.05, "resistivity": 55}[curve_key]
    scale = {"gamma": 1, "density": 0.01, "resistivity": 3}[curve_key]
    return round(max(0, base + wiggle * scale), 3)


def lithology_at_depth(intervals: list[dict], depth: float) -> dict | None:
    for interval in intervals:
        if interval["fromDepth"] <= depth <= interval["toDepth"]:
            return interval
    return None


def import_normalized_dataset(
    db: Session,
    dataset: dict,
    *,
    code_suffix: str = "",
    include_synthetic_curves: bool = True,
) -> Borehole:
    project = db.scalar(select(Project).where(Project.code == "DEMO-COAL"))
    if project is None:
        project = Project(code="DEMO-COAL", name="Demo Coal Block")
        db.add(project)
        db.flush()

    site_code = dataset["borehole"].get("block") or dataset["borehole"]["id"].split("-")[0]
    site = db.scalar(select(Site).where(Site.project_id == project.id).where(Site.code == site_code))
    if site is None:
        site = Site(project=project, code=site_code, name=site_code)
        db.add(site)
        db.flush()

    code = f"{dataset['borehole']['id']}{code_suffix}"
    existing = db.scalar(select(Borehole).where(Borehole.code == code))
    if existing is not None:
        return existing

    borehole = Borehole(
        site=site,
        code=code,
        title=dataset["borehole"]["title"],
        state=dataset["borehole"].get("state"),
        total_depth=dataset["borehole"]["totalDepth"],
        closure_note=dataset["borehole"].get("closureNote"),
        source_workbook=dataset["borehole"].get("sourceWorkbook"),
        source_sheet=dataset["borehole"].get("sourceSheet"),
        workflow_status=dataset["borehole"].get("workflowStatus") or "imported_for_review",
    )
    for item in dataset["lithologyIntervals"]:
        borehole.lithology_intervals.append(
            LithologyInterval(
                id=f"{code.lower()}-{item['id']}",
                source_row=item.get("sourceRow"),
                from_depth=item["fromDepth"],
                to_depth=item["toDepth"],
                lithology_code=item["lithologyCode"],
                lithology_label=item["lithologyLabel"],
                display_color=item.get("displayColor"),
                logged_color=item.get("loggedColor"),
                seam_name=item.get("seamName"),
                recovery=item.get("recovery"),
                recovery_percent=round((item["recovery"] / item["thickness"]) * 100, 2)
                if item.get("recovery") is not None and item.get("thickness")
                else None,
                rqd=item.get("rqd"),
                structural_features=item.get("structuralFeatures"),
                remark=item.get("remark"),
            )
        )
    for item in dataset["seamIntervals"]:
        borehole.seam_intervals.append(
            SeamInterval(
                id=f"{code.lower()}-{item['id']}",
                source_row=item.get("sourceRow"),
                name=item["name"],
                from_depth=item["fromDepth"],
                to_depth=item["toDepth"],
                thickness=item.get("thickness"),
                lithology_code=item.get("lithologyCode"),
                lithology_label=item.get("lithologyLabel"),
            )
        )

    if include_synthetic_curves:
        curve_defs = [
            ("gamma", "Gamma", "API", "#d97706"),
            ("density", "Density", "g/cc", "#059669"),
            ("resistivity", "Resistivity", "ohm.m", "#2563eb"),
        ]
        depth = 0.0
        samples_by_curve = {key: [] for key, *_ in curve_defs}
        while depth <= borehole.total_depth:
            interval = lithology_at_depth(dataset["lithologyIntervals"], depth)
            lithology_code = interval["lithologyCode"] if interval else None
            for key, *_ in curve_defs:
                samples_by_curve[key].append(
                    CurveSample(depth=round(depth, 2), value=synthetic_curve_value(key, depth, lithology_code))
                )
            depth = round(depth + 0.5, 3)
        for key, label, unit, color in curve_defs:
            borehole.curves.append(
                Curve(
                    key=key,
                    label=label,
                    unit=unit,
                    color=color,
                    source_type="synthetic_from_excel_profile",
                    samples=samples_by_curve[key],
                )
            )

    borehole.display_layouts.append(
        DisplayLayout(name="Imported Excel Review", mode="runtime", settings=default_borehole_layout())
    )
    borehole.source_imports.append(
        SourceImport(
            import_type="excel",
            source_name=dataset["borehole"].get("sourceWorkbook") or "Excel workbook",
            status="parsed",
            summary=dataset["profile"],
        )
    )
    db.add(borehole)
    db.flush()
    replace_validation_issues(borehole, validate_borehole(borehole))
    db.commit()
    db.refresh(borehole)
    return borehole


def import_excel_workbook(
    db: Session,
    workbook_path: Path,
    *,
    code_suffix: str = "",
    include_synthetic_curves: bool = True,
) -> Borehole:
    dataset = normalize_excel_workbook(workbook_path)
    return import_normalized_dataset(
        db,
        dataset,
        code_suffix=code_suffix,
        include_synthetic_curves=include_synthetic_curves,
    )
